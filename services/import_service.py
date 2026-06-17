from __future__ import annotations
import logging
import unicodedata
from datetime import datetime, date
from pathlib import Path

from db.database import get_conn

logger = logging.getLogger(__name__)

PAYMENT_MAP = {
    "DINHEIRO": "CASH",
    "CASH": "CASH",
    "PIX": "PIX",
    "CARTAO": "CARD",
    "CARTÃO": "CARD",
    "CARD": "CARD",
    "DEBITO": "CARD_DEBIT",
    "DÉBITO": "CARD_DEBIT",
    "CARD_DEBIT": "CARD_DEBIT",
    "CARTAO_DEBITO": "CARD_DEBIT",
    "CARTÃO_DÉBITO": "CARD_DEBIT",
    "CREDITO": "CARD_CREDIT",
    "CRÉDITO": "CARD_CREDIT",
    "CARD_CREDIT": "CARD_CREDIT",
    "CARTAO_CREDITO": "CARD_CREDIT",
    "CARTÃO_CRÉDITO": "CARD_CREDIT",
}


def _normalize(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    ).replace(" ", "_")


def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_time(val) -> str:
    if val is None:
        return "12:00"
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    s = str(val).strip()
    try:
        parts = s.split(":")
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    except Exception:
        return "12:00"


def _parse_cents(val) -> int:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(round(float(val) * 100))
    s = str(val).strip().replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        if s.rindex(",") > s.rindex("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return int(round(float(s) * 100))
    except ValueError:
        return 0


def _parse_qty(val) -> int:
    if val is None:
        return 1
    try:
        return max(1, int(float(str(val).replace(",", "."))))
    except (ValueError, TypeError):
        return 1


def _detect_columns(header: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}

    aliases: dict[str, list[str]] = {
        "data": ["data", "date", "dt", "dia"],
        "hora": ["hora", "time", "horario", "hr"],
        "pedido": ["pedido", "order", "venda", "nro", "num", "numero", "numero_pedido", "order_id", "cod", "codigo"],
        "descricao": ["descricao", "item", "produto", "product", "description", "desc", "nome", "servico"],
        "quantidade": ["quantidade", "qtd", "qty", "qnt", "qte"],
        "preco_unitario": [
            "preco_unitario", "preco_unit", "preco", "valor", "valor_unitario",
            "unit_price", "vl_unitario", "vl_unit", "preco unitario", "valor unitario",
        ],
        "pagamento": [
            "pagamento", "payment", "forma_pagamento", "forma", "metodo",
            "metodo_pagamento", "forma pagamento",
        ],
    }

    for i, raw in enumerate(header):
        h = _normalize(str(raw or "").strip())
        for field, names in aliases.items():
            if field not in col_map and h in names:
                col_map[field] = i
                break

    return col_map


def parse_excel(path: str) -> tuple[list[dict], list[str]]:
    """Parse Excel file. Returns (records, warnings)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl não instalado. Execute: pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Planilha vazia.")

    header = [str(c or "") for c in rows[0]]
    col_map = _detect_columns(header)

    required = ["data", "descricao", "preco_unitario"]
    missing = [k for k in required if k not in col_map]
    if missing:
        raise ValueError(
            f"Colunas obrigatórias não encontradas: {', '.join(missing)}.\n"
            "A planilha deve ter cabeçalho com: data, descricao, preco_unitario.\n"
            "Use o botão 'Baixar Modelo' para ver o formato esperado."
        )

    records = []
    warnings = []

    for row_idx, row in enumerate(rows[1:], start=2):
        data_val = row[col_map["data"]]
        if data_val is None:
            continue
        dt = _parse_date(data_val)
        if not dt:
            warnings.append(f"Linha {row_idx}: data inválida '{data_val}', ignorada.")
            continue

        hora = _parse_time(row[col_map["hora"]]) if "hora" in col_map else "12:00"
        pedido_key = str(row[col_map["pedido"]] or "").strip() if "pedido" in col_map else ""
        descricao = str(row[col_map["descricao"]] or "").strip()
        if not descricao:
            warnings.append(f"Linha {row_idx}: descrição vazia, ignorada.")
            continue

        preco = _parse_cents(row[col_map["preco_unitario"]])
        if preco <= 0:
            warnings.append(f"Linha {row_idx}: preço inválido '{row[col_map['preco_unitario']]}', ignorada.")
            continue

        quantidade = _parse_qty(row[col_map["quantidade"]] if "quantidade" in col_map else None)
        pagamento_raw = str(row[col_map["pagamento"]] or "").strip().upper() if "pagamento" in col_map else "PIX"
        pagamento = PAYMENT_MAP.get(pagamento_raw, "PIX")

        records.append({
            "data": dt,
            "hora": hora,
            "pedido_key": pedido_key,
            "descricao": descricao,
            "quantidade": quantidade,
            "preco_unitario": preco,
            "pagamento": pagamento,
        })

    return records, warnings


def _group_orders(records: list[dict]) -> list[dict]:
    from collections import OrderedDict

    groups: dict = OrderedDict()

    for i, rec in enumerate(records):
        key = (rec["data"], rec["pedido_key"]) if rec["pedido_key"] else (rec["data"], f"_auto_{i}")

        if key not in groups:
            groups[key] = {
                "data": rec["data"],
                "hora": rec["hora"],
                "items": [],
                "pagamento": rec["pagamento"],
            }
        groups[key]["items"].append({
            "descricao": rec["descricao"],
            "quantidade": rec["quantidade"],
            "preco_unitario": rec["preco_unitario"],
        })
        if rec["pagamento"]:
            groups[key]["pagamento"] = rec["pagamento"]

    return list(groups.values())


def import_from_excel(path: str) -> dict:
    """Import historical sales from Excel into the database.

    Returns: {imported_orders, skipped, errors, warnings}
    """
    records, warnings = parse_excel(path)
    if not records:
        return {"imported_orders": 0, "skipped": 0, "errors": [], "warnings": warnings}

    orders = _group_orders(records)
    imported = 0
    errors = []

    with get_conn() as conn:
        date_sessions: dict[str, int] = {}

        for order in orders:
            dt = order["data"]
            if dt not in date_sessions:
                r = conn.execute(
                    """
                    INSERT INTO cash_sessions(opened_at, closed_at, opening_cash_cents,
                                             closing_cash_cents, status, notes)
                    VALUES (?, ?, 0, 0, 'CLOSED', 'Importação retroativa Excel')
                    """,
                    (f"{dt} 08:00:00", f"{dt} 23:59:59"),
                )
                date_sessions[dt] = int(r.lastrowid)

            session_id = date_sessions[dt]
            created_at = f"{dt} {order['hora']}:00"

            total_cents = sum(
                item["quantidade"] * item["preco_unitario"]
                for item in order["items"]
            )

            try:
                r = conn.execute(
                    """
                    INSERT INTO orders(table_id, customer_id, status, total_cents, discount_cents,
                                       created_at, closed_at, cash_session_id, fiscal_status)
                    VALUES (NULL, NULL, 'CLOSED', ?, 0, ?, ?, ?, 'PENDING')
                    """,
                    (total_cents, created_at, created_at, session_id),
                )
                order_id = int(r.lastrowid)

                for item in order["items"]:
                    conn.execute(
                        """
                        INSERT INTO order_items(order_id, product_id, qty, unit_price_cents, notes)
                        VALUES (?, NULL, ?, ?, ?)
                        """,
                        (order_id, item["quantidade"], item["preco_unitario"], item["descricao"]),
                    )

                conn.execute(
                    """
                    INSERT INTO payments(order_id, method, amount_cents, change_cents)
                    VALUES (?, ?, ?, 0)
                    """,
                    (order_id, order["pagamento"], total_cents),
                )

                imported += 1
            except Exception as e:
                errors.append(f"Pedido {dt} {order['hora']}: {e}")

        conn.commit()

    try:
        from services.external_retry_service import enqueue_publish_report_retry
        from services.settings_service import is_report_auto_publish_enabled
        enqueue_publish_report_retry(auto_publish=bool(is_report_auto_publish_enabled()))
    except Exception:
        logger.exception("Failed to queue publish retry after import.")

    return {
        "imported_orders": imported,
        "skipped": 0,
        "errors": errors,
        "warnings": warnings,
    }


def generate_template_excel(path: str) -> None:
    """Generate a sample Excel template."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl não instalado.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendas"

    headers = ["data", "hora", "pedido", "descricao", "quantidade", "preco_unitario", "pagamento"]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    samples = [
        ["01/01/2026", "10:30", "",  "Sorvete de Chocolate", 2, "5,50",  "DINHEIRO"],
        ["01/01/2026", "14:00", "2", "Açaí 300ml",           1, "12,00", "PIX"],
        ["01/01/2026", "14:00", "2", "Casquinha Simples",    2, "4,00",  "PIX"],
        ["02/01/2026", "09:15", "",  "Milk Shake",           1, "15,00", "CARTAO"],
    ]
    for row_idx, row in enumerate(samples, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col, w in enumerate([14, 8, 10, 30, 12, 16, 14], start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    ws2 = wb.create_sheet("Instruções")
    lines = [
        "INSTRUÇÕES DE PREENCHIMENTO",
        "",
        "Colunas OBRIGATÓRIAS:",
        "  data          - Data da venda (DD/MM/AAAA ou AAAA-MM-DD)",
        "  descricao     - Descrição do item vendido",
        "  preco_unitario - Preço unitário  ex: 5,50  ou  5.50",
        "",
        "Colunas opcionais:",
        "  hora    - Horário da venda (HH:MM). Padrão: 12:00",
        "  pedido  - Número do pedido. Itens com mesmo pedido+data viram UMA venda.",
        "            Deixe em branco para cada linha ser uma venda separada.",
        "  quantidade  - Quantidade vendida. Padrão: 1",
        "  pagamento   - Forma de pagamento (ver abaixo). Padrão: PIX",
        "",
        "Formas de pagamento aceitas:",
        "  DINHEIRO  •  PIX  •  CARTAO  •  DEBITO  •  CREDITO",
        "",
        "Exemplo de agrupamento (pedido '2' tem 2 itens numa venda só):",
        "  01/01/2026 | 14:00 | 2 | Açaí 300ml       | 1 | 12,00 | PIX",
        "  01/01/2026 | 14:00 | 2 | Casquinha Simples | 2 |  4,00 | PIX",
        "",
        "ATENÇÃO: Não importe o mesmo arquivo mais de uma vez.",
    ]
    for i, line in enumerate(lines, start=1):
        ws2.cell(row=i, column=1, value=line)
    ws2.column_dimensions["A"].width = 72

    wb.save(path)
